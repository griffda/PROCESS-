#!/usr/bin/env python

"""
  Append summary data to a spreadsheet data_summary.xlsx or name specified.

  Michael Kovari

    + 31/07/2017: Initial version created
"""

import os
import argparse
import process_io_lib.mfile as mf
from process_io_lib.mfile import make_plot_dat
from openpyxl import Workbook, styles, load_workbook
from openpyxl.styles import  Font, Border, Side


def append_line(spreadsheet, custom_keys, mfile_data):
    """Function to add data to a spreadsheet using MFILE"""
    try:
        num_scans = int(mfile_data.data["isweep"].get_scan(-1))
    except KeyError:
        num_scans = 1
    print('num_scans', num_scans)

    val_keys = []
    mfile_keys = mfile_data.data.keys()

    try:
        wb = load_workbook(spreadsheet)
    except:
        wb = Workbook()

    # grab the active worksheet
    if args.s:
        ws = wb.create_sheet()
        print('New sheet created:',args.s)
    else:
        ws = wb.active

    var_descriptions = ['']
    var_names = ['']
    for key in custom_keys:
        if key in mfile_keys:
            var_description = mfile_data.data[key].var_description.replace(" ", "_")
            var_descriptions = var_descriptions + [var_description]

            val_keys.append(key)
            var_names = var_names + [key]

    # Print the descriptions only once on each sheet
    if args.s == False:
        if ws.max_row == 1:
            ws.append(var_descriptions)
    else:
        ws.append(var_descriptions)


    ws.append(var_names)

    # Write rows of values. One row for each scan.
    for num in range(num_scans):
         new_row = ['']
         value = ""
         for vkey in val_keys:
             if vkey in header_variables:
                 # These are found only once at the top of the MFILE.
                 value = mfile_data.data[vkey].get_scan(-1)
             else:
                # These are present in the MFILE for each scan point
                value = mfile_data.data[vkey].get_scan(num+1)
             new_row = new_row + [value]

         ws.append(new_row)

    # Save the spreadsheet
    wb.save(spreadsheet)

#-----------------------------------------------------------------------------


if __name__ == "__main__":

    # Setup command line arguments
    parser = argparse.ArgumentParser(description='Append summary data to a spreadsheet data_summary.xlsx or name specified.')

    parser.add_argument('-p', metavar='p', type=str, nargs='+',
                        help='add new variables to the output')

    parser.add_argument('-f', metavar='f', type=str, help='File to read as MFILE.DAT')

    parser.add_argument('-x', metavar='x', type=str,
                        help='Workbook (.xlsx) file to append to')

    parser.add_argument('-s', metavar='s', type=str,
                        help='Start a new worksheet (tab) with specified name')

    parser.add_argument("--defaults", help="run with default params", action="store_true")

    parser.add_argument("--reset-config", help="Reset xls.conf", action="store_true")

    args = parser.parse_args()

    default_variables = ['runtitle','username','date','iscan', 'rmajor', 'aspect', 'powfmw']
    header_variables = ['procver','date','time','username','runtitle','tagno','isweep','nsweep']

    # If user has specified an MFILE file that isn't MFILE.DAT pass the filename to
    # MFILE() class.
    if args.f:
        M = mf.MFile(filename=args.f)
    else:
        M = mf.MFile()

    # If user has specified a workbook file that isn't data_summary.xlsx
    if args.x:
        spreadsheet = args.x
    else:
        spreadsheet = 'data_summary.xlsx'

    print('spreadsheet name', spreadsheet)

    # Get files in current directory to check for the config file.
    current_directory = os.listdir(".")
    if "xls.conf" not in current_directory or args.reset_config:
        print('Configuration file xls.conf not found in the local directory')
        conf_file = open("xls.conf", "a")
        for item in default_variables:
            conf_file.write(item + "\n")
        conf_file.close()
        print('A default configuration file xls.conf has been written ')

    # Read the config file.
    INPUT_CONFIG = mf.read_mplot_conf("xls.conf")

    # If the user added new parameters in the command line add them to the
    # INPUT_CONFIG list to pass to make_plot_dat()
    if args.p:
        conf_file = open("xls.conf", "a")
        for item in args.p:
            if item not in INPUT_CONFIG:
                conf_file.write(item + "\n")
        conf_file.close()

    INPUT_CONFIG = mf.read_mplot_conf("xls.conf")
    print('Variables to be output:')
    print(INPUT_CONFIG)

    if args.defaults:
        append_line(spreadsheet, INPUT_CONFIG, M)
    else:
        append_line(spreadsheet, INPUT_CONFIG, M)
